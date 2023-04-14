from flask import Flask, render_template, request, redirect, session, url_for, g, abort, flash
import sqlite3, random, string, os
from openpyxl import load_workbook
from werkzeug.utils import secure_filename
from werkzeug.datastructures import FileStorage
from io import BytesIO

UPLOAD_FOLDER = os.path.join(os.path.abspath(os.path.dirname(__file__)), 'uploads')
app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.secret_key = 'turnkeytechnologies'
@app.before_request
def before_request():
    g.user = None

    if 'user_id' in session:
        user = [x for x in credentials if x.user_id == session['user_id']][0]
        g.user = user

class Survey:
    def __init__(self, name, questions):
        self.name = name
        self.questions = questions

class Student:
    def __init__(self, id, surname, title, given_names, teaching_period, unit_code, team_id):
        self.id = id
        self.surname = surname
        self.title = title
        self.given_names = given_names
        self.teaching_period = teaching_period
        self.unit_code = unit_code
        self.team_id = team_id
        self.spe1_responses = {}
        self.spe2_responses = {}
        self.spe1_scores = []
        self.spe2_scores = []
        self.spe1_total = 0
        self.spe2_total = 0
        self.accumulated_spe = 0

    def get_student(self):
        return self.id 

    def get_unit(self):
        return self.unit_code
    
    def get_id(self):
        return self.id

    def get_team(self, teams):
        for team in teams:
            if team.id == self.team_id:
                return team
        return None
    
    def calc_spe1(self):
        if self.spe1_scores:
            self.spe1_total = round(sum(self.spe1_scores)*(2.5/75), 1)
            return self.spe1_total
    
    def calc_spe2(self):
        if self.spe2_scores:
            self.spe2_total = round(sum(self.spe2_scores)*(2.5/75), 1)
            return self.spe2_total
    
    def calc_total_spe(self):
        self.accumulated_spe = round(self.spe1_total + self.spe2_total, 1)
        return self.accumulated_spe
    

class UnitCoordinator:
    def __init__(self, id, surname, title, given_names, unit_codes, email):
        self.id = id
        self.surname = surname
        self.title = title
        self.given_names = given_names
        self.unit_codes = unit_codes
        self.email = email
        self.surveys = []

        self.student_conn = sqlite3.connect('students.db')
        self.student_cur = self.student_conn.cursor()

    def upload_students_from_excel(self, uploaded_file):
        if not isinstance(uploaded_file, FileStorage):
            # `uploaded_file` is a file path, so open the file and create a file object
            with open(uploaded_file, 'rb') as f:
                filecontents = f.read()
            uploaded_file = FileStorage(stream=BytesIO(filecontents), filename=os.path.basename(uploaded_file))

        if not uploaded_file.filename.endswith('.xlsx'):
            flash('Invalid file type. Please upload an Excel file.')
            return redirect(url_for('uc_dashboard'))

        filename = secure_filename(uploaded_file.filename)
        filecontents = uploaded_file.stream.read()
        uploaded_file.stream.seek(0)  # Reset the file pointer to the beginning of the file

        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        with open(filepath, 'wb') as f:
            f.write(filecontents)

        try:
            wb = load_workbook(filename=filepath, read_only=True)
        except Exception as e:
            flash(f"Failed to load workbook: {str(e)}")
            return redirect(url_for('uc_dashboard'))

        ws = wb.active
        count = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            id, surname, title, given_names, teaching_period, unit_code, team_id = row
            student = admin.create_student(id, surname, title, given_names, teaching_period, unit_code, team_id)
            count += 1

        message = f"{count} student(s) have been uploaded and saved in the database."
        flash(message)
        return redirect(url_for('uc_dashboard'))
        
    def create_survey(self, name, questions):
        survey = Survey(name, questions)
        self.surveys.append(survey)
        return Survey(name, questions)
    
class Unit:
    def __init__(self, unit_code, name, teaching_period):
        self.unit_code = unit_code
        self.name = name
        self.teaching_period = teaching_period

    @staticmethod
    def get_unit(unit_code):
        for unit in admin.units:
            if unit.unit_code == unit_code:
                return unit
        return None

class Team:
    def __init__(self, id, unit_code):
        self.id = id
        self.unit_code = unit_code
        self.members = []

    def add_member(self, student):
        self.members.append(student)
    
    def remove_member(self, student):
        if student in self.members:
            self.members.remove(student)
    
    def get_members(self):
        return self.members
    
    def has_member(self, student):
        return student in self.members    

class Admin:
    def __init__(self, id):
        self.id = id
        self.units = []
        self.students = []
        self.teams = []
        self.enrollments = []

        self.student_conn = sqlite3.connect('students.db')
        self.student_cur = self.student_conn.cursor()
        self.student_cur.execute('''CREATE TABLE IF NOT EXISTS students
                                    (id TEXT, surname TEXT, title TEXT, given_names TEXT,
                                    teaching_period TEXT, unit_code TEXT, team_id TEXT)''')

        self.uc_conn = sqlite3.connect('uc.db')
        self.uc_cur = self.uc_conn.cursor()
        self.uc_cur.execute('''CREATE TABLE IF NOT EXISTS unitcoordinators
                                (id TEXT, surname TEXT, title TEXT, given_names TEXT,
                                unit_codes TEXT, email TEXT)''')
    def get_admin_id(self):
        return self.id
    
    def create_student(self, id, surname, title, given_names, teaching_period, unit_code, team_id):
        student = Student(id, surname, title, given_names, teaching_period, unit_code, team_id)
        self.student_cur.execute("INSERT INTO students VALUES (?, ?, ?, ?, ?, ?, ?)",
                                 (student.id, student.surname, student.title, student.given_names,
                                  student.teaching_period, student.unit_code, student.team_id))
        self.student_conn.commit()
        self.students.append(student)
        return student

    def delete_student(self, student):
        self.student_cur.execute("DELETE FROM students WHERE id=?", (student.id,))
        self.student_conn.commit()

    def get_student_by_id(self, student_id):
        for student in self.students:
            if student.id == student_id:
                return student
        return None

    def create_uc(self, id, surname, title, given_names, unit_codes, email):
        uc = UnitCoordinator(id, surname, title, given_names, unit_codes, email)
        self.uc_cur.execute("INSERT INTO unitcoordinators VALUES (?, ?, ?, ?, ?, ?)", 
                             (uc.id, uc.surname, uc.title, uc.given_names, ",".join(uc.unit_codes), uc.email))
        self.uc_conn.commit()
        return uc

    def delete_uc(self, uc):
        self.uc_cur.execute("DELETE FROM unitcoordinators WHERE id=?", (uc.id,))
        self.uc_conn.commit()

    def create_unit(self, unit_code, name, teaching_period):
        unit = Unit(unit_code, name, teaching_period)
        self.units.append(unit)

    def get_unit(self, unit_code):
        for unit in self.units:
            if unit.unit_code == unit_code:
                return unit
        return None
    
    def create_team(self, id, unit_code):
        team = Team(id, unit_code)
        self.teams.append(team)
        return team

    def enroll_students(self):
        for team in self.teams:
            for student in self.students:
                if student.unit_code == team.unit_code and student.team_id == team.id:
                    team.add_member(student)

    def print_enrollments(self):
        for team in self.teams:
            print(f"Team {team.id} ({team.unit_code}):")
            if len(team.members) == 0:
                print("No members")
            else:
                for member in team.members:
                    print(f"Student ID: {member.id}, Unit Code: {member.unit_code}")
            print()  
        
admin = Admin('110001')
print()
print("Created Admin:")
print(admin.id)
print()

admin.create_unit('ICT302', 'IT Professional Practice Project', 'TJD23'),
admin.create_unit('ICT201', 'Information Technology Project Management', 'TJD23')

print("Created Units:")
for unit in admin.units:
    print(unit.unit_code, unit.name, unit.teaching_period)
print()

admin.create_team('GR1', 'ICT302')
admin.create_team('GR2', 'ICT302')
admin.create_team('GR1', 'ICT201')
admin.create_team('GR2', 'ICT201')

print("Created Teams:")
for team in admin.teams:
    print(team.id, team.unit_code)
print()

class Credentials:
    def __init__(self, user_id):
        self.user_id = user_id
        self.password = self.generate_password()
    
    def generate_password(self):
        # Generate a random 4-character password
        return ''.join(random.choices(string.ascii_uppercase + string.digits, k=4))
    
    def __repr__(self):
        return f'<Credentials: {self.user_id}>'

uc = admin.create_uc('220001', 'Parry', 'Dr', 'David', ['ICT302', 'ICT201'], 'david.parry@murdoch.edu.au')
print("Created unit coordinator:")
print(uc.id, uc.surname, uc.title, uc.given_names, uc.unit_codes, uc.email)
print()

stds = [
    Student('330001', "Tiana", "Miss", "Anjara", "TJD23", "ICT302", 'GR1'),
    Student('330002', "Emadaldin", "Mr", "Ahmed", "TJD23", "ICT302", 'GR1'),
    Student('330003', "Shaikh", "Miss", "Fatima", "TJD23", "ICT302", 'GR1'),
    Student('330004', "Srinivas", "Miss", "Sanhita", "TJD23", "ICT302", 'GR2'),
    Student('330005', "Doe", "Miss", "Jane", "TJD23", "ICT302", 'GR2'),
    Student('330006', "Williams", "Mr", "Peter", "TJD23", "ICT302", 'GR2'),
    Student('330007', "Patterson", "Miss", "Noor", "TJD23", "ICT201", 'GR1'),
    Student('330008', "Bensaid", "Miss", "Anais", "TJD23", "ICT201", 'GR1'),
    Student('330009', "Doe", "Mr", "John", "TJD23", "ICT201", 'GR1'),
    Student('330010', "Ivanovic", "Miss", "Ana", "TJD23", "ICT201", 'GR2'),
    Student('330011', "Richards", "Mr", "Steven", "TJD23", "ICT201", 'GR2'),
    Student('330012', "Abdullah", "Miss", "Brenda", "TJD23", "ICT201", 'GR2')
]

for std in stds:
    admin.create_student(
        id=std.id,
        surname=std.surname,
        title=std.title,
        given_names=std.given_names,
        teaching_period=std.teaching_period,
        unit_code=std.unit_code,
        team_id=std.team_id
    )

# Get all existing object IDs and create credentials for them
ids = []
ids.append(admin.id)
ids.append(uc.id)
print("Created Students:")
for student in admin.students:
    ids.append(student.id)
    print(student.id, student.surname, student.title, student.given_names, student.teaching_period, student.unit_code, student.team_id)
print()

credentials = []
for user_id in ids:
    cred = Credentials(user_id)
    credentials.append(cred)

admin.enroll_students()
admin.print_enrollments()

# Print out all credentials objects
print("User Credentials:")
for cred in credentials:
    print(f'ID: {cred.user_id}, Password: {cred.password}')
print()

SE_questions = ["Provide a brief description of how you believe you contributed to the project process for the duration of the trimester: ",
                "What skills and knowledge do you now know you need to develop for your future work in the IT industry and/or what issues of your own working style do you need to address?"
]
PE_questions = ["Rate the work and effort this team member contributed towards the project documents:",
    "Rate this team member's willingness to work and to take responsibility:",
    "Rate this team member's level of communication and participation within the group and during meetings:",
    "Rate this team member's contribution towards the management of the project, e.g. work delivered on time:",
    "Rate this team member's problem-solving and creativity on behalf of the group work:",
    "Provide a brief description of how you believe this team member contributed to the project process for the duration of the trimester:"]

surveys = [
    Survey("SE1", SE_questions),
    Survey("SE2", SE_questions),
    Survey("PE1.1", SE_questions),
    Survey("PE1.2", SE_questions),
    Survey("PE2.1", SE_questions),
    Survey("PE2.2", SE_questions)
]

for survey in surveys:
    uc.create_survey(
        name = survey.name,
        questions = survey.questions
    )
print()

'''
print("SPE 1 Responses: ")
for std in admin.students:
    print(std.id, std.spe1_responses)

print()
print("SPE 2 Responses: ")
for std in admin.students:
    print(std.id, std.spe2_responses)
print()
'''
admin.students[0].spe1_scores = [5,2,4,2,4,5,3,3,4,2,2,2,1,4,5]
admin.students[0].spe2_scores = [5,2,4,1,3,4,5,2,4,2,2,2,1,4,5]
admin.students[1].spe1_scores = [5,5,4,3,3,2,2,4,5,2,3,4,1,5,2]
admin.students[1].spe2_scores = [5,4,5,4,5,2,3,4,5,5,2,4,1,3,2]
admin.students[2].spe1_scores = [4,3,1,5,3,2,4,1,4,4,5,3,1,2,4]
admin.students[2].spe2_scores = [4,4,3,2,4,4,5,5,5,5,2,2,2,4,4]
admin.students[3].spe2_scores = [5,5,5,5,4,4,5,5,5,5,5,5,5,5,5]
admin.students[3].spe1_scores = [4,2,5,1,3,4,2,3,5,3,1,3,3,3,4]
admin.students[4].spe2_scores = [5,5,4,4,4,4,2,4,2,4,2,4,2,1,1]
admin.students[4].spe1_scores = [1,2,3,4,4,4,5,2,2,4,2,2,4,2,4]
admin.students[5].spe2_scores = [5,3,5,4,2,2,3,2,3,4,1,3,1,3,3]
admin.students[5].spe1_scores = [5,2,4,2,4,5,3,3,4,2,2,5,5,5,5]
admin.students[6].spe2_scores = [5,2,4,3,3,1,3,1,4,2,2,2,1,4,5]
admin.students[6].spe1_scores = [4,4,4,2,4,4,5,2,5,5,2,2,1,3,4]
admin.students[7].spe2_scores = [5,5,5,4,4,4,4,4,4,5,3,3,4,5,5]
admin.students[7].spe1_scores = [4,4,4,4,4,4,4,4,4,4,4,4,4,4,4]
admin.students[8].spe2_scores = [3,3,3,3,3,5,5,2,3,3,1,1,3,3,4]
admin.students[8].spe1_scores = [4,5,4,5,5,5,4,4,4,5,4,5,4,4,5]
admin.students[8].spe2_scores = [3,5,4,2,3,1,3,4,5,3,2,4,2,1,3]
admin.students[9].spe1_scores = [5,4,2,2,4,1,3,5,3,2,3,1,3,3,2]
admin.students[9].spe2_scores = [2,3,2,1,2,3,2,4,5,3,2,3,4,5,1]
admin.students[10].spe1_scores = [1,1,2,2,3,2,1,2,1,2,1,2,3,3,3]
admin.students[10].spe2_scores = [2,2,1,1,3,1,3,1,1,3,4,2,2,3,3]
admin.students[11].spe1_scores = [5,4,2,5,5,5,5,3,4,2,2,2,1,4,5]
admin.students[11].spe2_scores = [5,2,4,1,3,4,3,5,5,5,2,3,5,4,5]

for std in admin.students:
    std.calc_spe1()
    std.calc_spe2()
    std.calc_total_spe()
    print(std.id, std.spe1_scores, std.spe2_scores, std.spe1_total, std.spe2_total, std.accumulated_spe)


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        session.pop('user_id', None)

        user_id = request.form['user_id']
        password = request.form['password']
        
        user = [x for x in credentials if x.user_id == user_id][0]
        if user and user.password == password:
            session['user_id'] = user.user_id
            if user_id.startswith('11'):
                return redirect(url_for('admin_dashboard'))
            elif user_id.startswith('22'):
                return redirect(url_for('uc_dashboard'))
            elif user_id.startswith('33'):
                return redirect(url_for('student_dashboard'))      
        return redirect(url_for('login'))

    return render_template('login.html')

@app.route('/student_dashboard')
def student_dashboard():
    if not g.user:
        return redirect(url_for('login'))

    # get the current user's ID
    user_id = session['user_id']
    
    # get the student object corresponding to the current user's ID
    current_student = admin.get_student_by_id(user_id)

    # get the current student's unit object
    current_unit = admin.get_unit(current_student.unit_code)

    return render_template('student_dashboard.html', current_unit=current_unit, current_student=current_student)

@app.route('/admin_dashboard')
def admin_dashboard():
    admin_id = admin.id
    creds = credentials
    return render_template('admin_dashboard.html', admin_id=admin_id, creds=creds)

@app.route('/uc_dashboard', methods=['GET', 'POST'])
def uc_dashboard():
    if request.method == 'POST':
        file = request.files['file']
        if file.filename != '':
            file_path = secure_filename(file.filename)
            file.save(file_path)
            conn = admin.student_conn
            uc.upload_students_from_excel(file_path)
            flash('File uploaded successfully', 'success')
            return redirect(url_for('index'))
                
    unitCoord = uc
    unitOne = admin.get_unit(uc.unit_codes[0])
    unitTwo = admin.get_unit(uc.unit_codes[1])
    return render_template('uc_dashboard.html', unitCoord=unitCoord, unitOne=unitOne, unitTwo=unitTwo)

@app.route('/portal/<unit_code>')
def portal(unit_code):
    # Get the unit object using the unit_code passed in the URL
    unit = admin.get_unit(unit_code)

    if not g.user:
        return redirect(url_for('login'))

    # get the current user's ID
    user_id = session['user_id']
    
    # get the student object corresponding to the current user's ID
    current_student = admin.get_student_by_id(user_id)

    # get the current student's unit object
    current_unit = admin.get_unit(current_student.unit_code)

    if current_unit.unit_code != unit_code:
        return redirect(url_for('student_dashboard'))

    if request.args.get('redirect') == 'incident_report':
        return redirect(url_for('incident_report'))
    
    # set up routing for SPE1 and SPE2 links
    if request.args.get('spe') == '1':
        return redirect(url_for('SPE1'))
    elif request.args.get('spe') == '2':
        return redirect(url_for('SPE2'))

    return render_template('portal.html', current_unit=unit, current_student=current_student)

@app.route('/SPE1', methods=['GET', 'POST'])
def SPE1():
    # get the current user's ID from the request args
    user_id = request.args.get('user_id')
    
    # get the student object corresponding to the current user's ID
    current_student = admin.get_student_by_id(user_id)
    
    if not current_student:
        # handle error if student not found
       print('Student not found')
        
    team = current_student.get_team(admin.teams)
    team_members = [s for s in team.get_members() if s != current_student]
    member1 = team_members[0]
    member2 = team_members[1]

    if request.method == 'POST':
        q1 = int(request.form['q1'])
        q2 = int(request.form['q2'])
        q3 = int(request.form['q3'])
        q4 = int(request.form['q4'])
        q5 = int(request.form['q5'])
        q6 = request.form['q6']
        q7 = request.form['q7']
        q8 = int(request.form['q8'])
        q9 = int(request.form['q9'])
        q10 = int(request.form['q10'])
        q11 = int(request.form['q11'])
        q12 = int(request.form['q12'])
        q13 = request.form['q13']
        q14 = int(request.form['q14'])
        q15 = int(request.form['q15'])
        q16 = int(request.form['q16'])
        q17 = int(request.form['q17'])
        q18 = int(request.form['q18'])
        q19 = request.form['q19']


        # Save responses to variables
        responses = {'id': user_id, 'q1': q1, 'q2': q2, 'q3': q3, 'q4': q4, 'q5': q5,
                     'q6': q6, 'q7': q7, 'q8': q8, 'q9': q9, 'q10': q10, 'q11': q11, 'q12': q12,
                     'q13': q13, 'q14': q14, 'q15': q15, 'q16': q16, 'q17': q17, 'q18': q18, 'q19': q19}
        
        '''
        current_student['spe1_responses'].update(responses)
        '''
        current_student.spe1_scores.extend([q1, q2, q3, q4, q5])
        member1.spe1_scores.extend([q8, q9, q10, q11, q12])
        member2.spe1_scores.extend([q14, q15, q16, q17, q18])

        return render_template('thank_you.html', responses=responses, current_student=current_student, member1=member1, member2=member2)
    else:
        return render_template('SPE1.html', current_student=current_student, team_members=team_members, member1=member1, member2=member2)
    
@app.route('/SPE2', methods=['GET', 'POST'])
def SPE2():
    # get the current user's ID from the request args
    user_id = request.args.get('user_id')
    
    # get the student object corresponding to the current user's ID
    current_student = admin.get_student_by_id(user_id)
    
    if not current_student:
        # handle error if student not found
       print('Student not found')
        
    team = current_student.get_team(admin.teams)
    team_members = [s for s in team.get_members() if s != current_student]
    member1 = team_members[0]
    member2 = team_members[1]

    if request.method == 'POST':
        q1 = int(request.form['q1'])
        q2 = int(request.form['q2'])
        q3 = int(request.form['q3'])
        q4 = int(request.form['q4'])
        q5 = int(request.form['q5'])
        q6 = request.form['q6']
        q7 = request.form['q7']
        q8 = int(request.form['q8'])
        q9 = int(request.form['q9'])
        q10 = int(request.form['q10'])
        q11 = int(request.form['q11'])
        q12 = int(request.form['q12'])
        q13 = request.form['q13']
        q14 = int(request.form['q14'])
        q15 = int(request.form['q15'])
        q16 = int(request.form['q16'])
        q17 = int(request.form['q17'])
        q18 = int(request.form['q18'])
        q19 = request.form['q19']


        # Save responses to variables
        responses = {'id': user_id, 'q1': q1, 'q2': q2, 'q3': q3, 'q4': q4, 'q5': q5,
                     'q6': q6, 'q7': q7, 'q8': q8, 'q9': q9, 'q10': q10, 'q11': q11, 'q12': q12,
                     'q13': q13, 'q14': q14, 'q15': q15, 'q16': q16, 'q17': q17, 'q18': q18, 'q19': q19}
        
        '''
        current_student['spe2_responses'].update(responses)
        '''
        current_student.spe2_scores.extend([q1, q2, q3, q4, q5])
        member1.spe2_scores.extend([q8, q9, q10, q11, q12])
        member2.spe2_scores.extend([q14, q15, q16, q17, q18])

        return render_template('thank_you.html', responses=responses, current_student=current_student, member1=member1, member2=member2)
    else:
        return render_template('SPE2.html', current_student=current_student, team_members=team_members, member1=member1, member2=member2)
    
@app.route("/incident_report")
def incident_report():
    return render_template("incident_report.html")

@app.route('/ICT302_result')
def ICT302_result():
    unitCoord = uc
    student1 = admin.students[0]
    student2 = admin.students[1]
    student3 = admin.students[2]
    student4 = admin.students[3]
    student5 = admin.students[4]
    student6 = admin.students[5]
    return render_template('ICT302_result.html', unitCoord=unitCoord, student1=student1, student2=student2, student3=student3, student4=student4, student5=student5, student6=student6)

@app.route('/ICT201_result')
def ICT201_result():
    unitCoord = uc
    student7 = admin.students[6]
    student8 = admin.students[7]
    student9 = admin.students[8]
    student10 = admin.students[9]
    student11 = admin.students[10]
    student12 = admin.students[11]
    return render_template('ICT201_result.html', unitCoord=unitCoord, student7=student7, student8=student8, student9=student9, student10=student10, student11=student11, student12=student12)
    
@app.route('/thank_you', methods=['GET'])
def thank_you():
    return render_template('thank_you.html')

@app.route('/')
def index():
    student_records = [
        ['<br>330001', 'Tiana', 'Miss', 'Anjara', 'TJD23', 'ICT302', 'GR1'],
        ['330002', 'Emadaldin', 'Mr', 'Ahmed', 'TJD23', 'ICT302', 'GR1'],
        ['330003', 'Shaikh', 'Miss', 'Fatima', 'TJD23', 'ICT302', 'GR1'],
        ['330004', 'Srinivas', 'Miss', 'Sanhita', 'TJD23', 'ICT302', 'GR2'],
        ['330005', 'Doe', 'Miss', 'Jane', 'TJD23', 'ICT302', 'GR2'],
        ['330006', 'Williams', 'Mr', 'Peter', 'TJD23', 'ICT302', 'GR2'],
        ['330007', 'Patterson', 'Miss', 'Noor', 'TJD23', 'ICT201', 'GR1'],
        ['330008', 'Bensaid', 'Miss', 'Anais', 'TJD23', 'ICT201', 'GR1'],
        ['330009', 'Doe', 'Mr', 'John', 'TJD23', 'ICT201', 'GR1'],
        ['330010', 'Ivanovic', 'Miss', 'Ana', 'TJD23', 'ICT201', 'GR2'],
        ['330011', 'Richards', 'Mr', 'Steven', 'TJD23', 'ICT201', 'GR2'],
        ['330012', 'Abdullah', 'Miss', 'Brenda', 'TJD23', 'ICT201', 'GR2']
    ]
    message = "The student file has successfully been uploaded. The uploaded student records are:\n"
    for record in student_records:
        message += ' '.join(record) + '<br>'
    return message

@app.errorhandler(404)
def page_not_found(error):
    return render_template('404.html'), 404

@app.errorhandler(500)
def internal_error(error):
    return render_template('500.html'), 500


if __name__ == '__main__':
    app.run(debug=True)
